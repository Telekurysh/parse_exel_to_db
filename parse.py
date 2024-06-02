from openpyxl import load_workbook
import logging
import difflib
from example import data_dict


def compare_strings(string1, string2):
    matcher = difflib.SequenceMatcher(None, string1, string2)

    differences = []
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == 'replace':
            differences.append(f"Замена: '{string1[i1:i2]}' на '{string2[j1:j2]}'")
        elif tag == 'delete':
            differences.append(f"Удаление: '{string1[i1:i2]}'")
        elif tag == 'insert':
            differences.append(f"Вставка: '{string2[j1:j2]}'")
        elif tag == 'equal':
            differences.append(f"Совпадение: '{string1[i1:i2]}'")

    return differences


logger = logging.getLogger('my_logger')
logger.setLevel(logging.DEBUG)
file_handler = logging.FileHandler('parser.log', encoding='utf-8')
file_handler.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)


def read_excel_file(file_path):
    try:
        workbook = load_workbook(filename=file_path)

        sheet = workbook.active

        data = []

        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        for i, value in enumerate(data[0], start=1):
            if data_dict.get(i) != value:
                sorted_values = [data_dict[key] for key in sorted(data_dict.keys())]
                result_string = ' '.join(sorted_values)
                logger.addHandler(file_handler)
                logger.error("Шаблон поменялся")
                logger.error('\n'.join(compare_strings(result_string, " ".join(data[0]))))
                return False
        return data

    except Exception as e:
        print(f"Ошибка при чтении файла Excel: {e}")
        return None


# file_path = "0 4.xlsx"  # Укажите путь к вашему файлу Excel
# excel_data = read_excel_file(file_path)
# if excel_data:
#     print("Данные из файла Excel:")
#     print(excel_data[0])
#     with open("example.txt", "w") as f:
#         f.write(" ".join(excel_data[0]))
# print(len(excel_data) - 1)
# for row in excel_data:
#     print(row)
